"""Ingestor for NCES CIP-to-SOC crosswalk (CIP 2020 x SOC 2018).

Downloads the official XLSX from NCES and lands the CIP-SOC pairing data
into the bronze zone as raw.cip_soc_crosswalk.
"""

import logging
from pathlib import Path
from typing import Any

import requests
from pyiceberg.schema import Schema
from pyiceberg.types import (
    DateType,
    NestedField,
    StringType,
    TimestampType,
)

from brightsmith.bronze.base_ingestor import BaseIngestor

logger = logging.getLogger(__name__)


class CipSocCrosswalkIngestor(BaseIngestor):
    """Ingests the NCES CIP-to-SOC crosswalk into the bronze zone.

    Data source: XLSX download from NCES containing many-to-many
    CIP 2020 to SOC 2018 pairings.

    Grain: cipcode x soc_code (one row per CIP-SOC pairing)

    Key considerations:
    - XLSX has multiple sheets; use "CIP-SOC" as the authoritative sheet
    - CIP codes may arrive as floats from openpyxl -- must format as XX.XXXX strings
    - SOC codes arrive as strings in XX-XXXX format
    - 194 "no match" rows (SOC 99-9999) are preserved in Bronze; filtered in Silver
    """

    DOWNLOAD_URL = (
        "https://nces.ed.gov/ipeds/cipcode/Files/CIP2020_SOC2018_Crosswalk.xlsx"
    )
    FALLBACK_PATH = "data/raw/xlsx_cache/CIP2020_SOC2018_Crosswalk.xlsx"
    USER_AGENT = "FutureProof/0.1 (jeff@hyenastudios.com)"

    # XLSX column names (from the CIP-SOC sheet)
    COLUMN_MAP: dict[str, str] = {
        "CIP2020Code": "cipcode",
        "CIP2020Title": "cip_title",
        "SOC2018Code": "soc_code",
        "SOC2018Title": "soc_title",
    }

    # Target sheet name in the XLSX workbook
    TARGET_SHEET = "CIP-SOC"

    def fetch(self, entities: dict, method: str, **kwargs) -> dict:
        """Download CIP-SOC crosswalk XLSX and parse the CIP-SOC sheet.

        Args:
            entities: {entity_id: label} dict from source config.
            method: Fetch method name.
            **kwargs: Supports ``xlsx_path`` for local file (used in tests).

        Returns:
            Dict mapping each entity_id to a list of row dicts.
        """
        xlsx_path = kwargs.get("xlsx_path")
        if xlsx_path is not None:
            rows = self._read_xlsx(Path(xlsx_path))
        else:
            rows = self._download_and_read()

        return {entity_id: rows for entity_id in entities}

    def _download_and_read(self) -> list[dict[str, Any]]:
        """Download the XLSX from NCES and return parsed rows."""
        headers = {"User-Agent": self.USER_AGENT}

        try:
            response = requests.get(
                self.DOWNLOAD_URL, headers=headers, allow_redirects=True, timeout=120
            )
            response.raise_for_status()

            import tempfile

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(response.content)
                tmp_path = Path(tmp.name)
            rows = self._read_xlsx(tmp_path)
            tmp_path.unlink(missing_ok=True)
            return rows
        except Exception:
            logger.warning(
                "Download failed, falling back to %s", self.FALLBACK_PATH
            )
            return self._read_xlsx(Path(self.FALLBACK_PATH))

    def _read_xlsx(self, path: Path) -> list[dict[str, Any]]:
        """Read the CIP-SOC sheet from the XLSX and return rows as dicts."""
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Try the target sheet name first, fall back to active sheet
        if self.TARGET_SHEET in wb.sheetnames:
            ws = wb[self.TARGET_SHEET]
        else:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        # Find the header row
        header_row = None
        for row in rows_iter:
            str_cells = [c for c in row if isinstance(c, str) and c.strip()]
            if len(str_cells) >= 4:
                header_row = [str(c).strip() if c else "" for c in row]
                break

        if header_row is None:
            wb.close()
            raise ValueError("Could not find header row in XLSX")

        # Map column indices to our canonical names
        col_mapping: dict[int, str] = {}
        for idx, header in enumerate(header_row):
            if header in self.COLUMN_MAP:
                col_mapping[idx] = self.COLUMN_MAP[header]

        if len(col_mapping) < 4:
            wb.close()
            raise ValueError(
                f"Could not map all 4 required columns. Found: {col_mapping}. "
                f"Headers: {header_row}"
            )

        data_rows: list[dict[str, Any]] = []
        for row in rows_iter:
            record: dict[str, Any] = {}
            for col_idx, canonical_name in col_mapping.items():
                record[canonical_name] = row[col_idx] if col_idx < len(row) else None
            # Skip rows where cipcode and soc_code are both None (blank rows)
            if record.get("cipcode") is None and record.get("soc_code") is None:
                continue
            data_rows.append(record)

        wb.close()
        logger.info("Parsed %d crosswalk rows from XLSX", len(data_rows))
        return data_rows

    def flatten(self, raw_data: Any, entity_id: str) -> list[dict]:
        """Flatten raw XLSX rows into Iceberg-ready dicts.

        Handles:
        - CIP code formatting: openpyxl may return floats (52.0201) or
          strings ("52.0201"). Must always produce XX.XXXX string format.
        - SOC code: keep as string XX-XXXX.
        - Strip whitespace from title fields.

        Args:
            raw_data: List of row dicts from fetch().
            entity_id: Logical entity identifier.

        Returns:
            List of flat dicts with lowercase keys matching the Iceberg schema.
            Does NOT add metadata fields -- the framework handles those.
        """
        flat_rows: list[dict] = []
        skipped = 0

        for raw_row in raw_data:
            cipcode = self._coerce_cipcode(raw_row.get("cipcode"))
            soc_code = self._coerce_soc(raw_row.get("soc_code"))

            if cipcode is None or soc_code is None:
                skipped += 1
                continue

            cip_title = self._coerce_string(raw_row.get("cip_title"))
            soc_title = self._coerce_string(raw_row.get("soc_title"))

            if cip_title is None or soc_title is None:
                skipped += 1
                continue

            record = {
                "cipcode": cipcode,
                "cip_title": cip_title,
                "soc_code": soc_code,
                "soc_title": soc_title,
            }
            flat_rows.append(record)

        if skipped:
            logger.warning("Skipped %d rows with null/invalid fields", skipped)
        return flat_rows

    @staticmethod
    def _coerce_cipcode(value: Any) -> str | None:
        """Coerce CIP code to XX.XXXX string format.

        Handles:
        - Float from openpyxl: 52.0201 -> "52.0201"
        - Integer from openpyxl: 1 -> "01.0000"
        - String already in format: "52.0201" -> "52.0201"
        - String without dot: "520201" -> attempt to parse
        """
        if value is None:
            return None

        if isinstance(value, (int, float)):
            # Float from openpyxl -- format as XX.XXXX
            # e.g. 52.0201 -> "52.0201", 1.0 -> "01.0000"
            formatted = f"{value:07.4f}"
            return formatted

        s = str(value).strip()
        if not s:
            return None

        # Already has a dot -- ensure proper zero-padding
        if "." in s:
            parts = s.split(".")
            family = parts[0].zfill(2)
            detail = parts[1].ljust(4, "0")[:4]
            return f"{family}.{detail}"

        return s

    @staticmethod
    def _coerce_soc(value: Any) -> str | None:
        """Coerce SOC code to string, preserving XX-XXXX format."""
        if value is None:
            return None
        s = str(value).strip()
        return s if s else None

    @staticmethod
    def _coerce_string(value: Any) -> str | None:
        """Coerce to stripped string or None."""
        if value is None:
            return None
        s = str(value).strip()
        return s if s else None

    def get_schema(self) -> Schema:
        """Define the Iceberg table schema for raw.cip_soc_crosswalk.

        Matches the fields returned by flatten(). Grain is cipcode x soc_code.
        """
        return Schema(
            NestedField(1, "cipcode", StringType(), required=True),
            NestedField(2, "cip_title", StringType(), required=True),
            NestedField(3, "soc_code", StringType(), required=True),
            NestedField(4, "soc_title", StringType(), required=True),
            NestedField(5, "ingested_at", TimestampType(), required=True),
            NestedField(6, "source_url", StringType(), required=True),
            NestedField(7, "source_method", StringType(), required=True),
            NestedField(8, "load_date", DateType(), required=True),
        )
